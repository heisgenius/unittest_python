import openpyxl
class doExcel:
    def do_excel(self,button,caseid):
        wb=openpyxl.load_workbook("readData.xlsx")
        sheet = wb["testData"]
        headers =[]
        for i in range(1,6):
            headers.append(sheet.cell(1,i).value)
        # print(headers)

        test_data=[]
        for i in range(2,6):
            sub_data ={}
            for j in range(1,6):
                # print(sheet.cell(i,j).value)
                sub_data[headers[j-1]]=sheet.cell(i,j).value
            test_data.append(sub_data)
        if button == 'on':
            final_data = test_data
        else:
            final_data = []
            for item in test_data:
                if item['caseid'] in caseid:
                    final_data.append(item)
        return final_data
    def writeBack(self,row,actualres,testres):
        wb = openpyxl.load_workbook("readData.xlsx")
        sheet = wb['testData']
        sheet.cell(row,6).value = actualres
        sheet.cell(row,7).value = testres
        wb.save('readData.xlsx')

#
if __name__ == '__main__':
    res =doExcel().do_excel('on',[1,2])
    print(res)